import random
import openpyxl

#extract
def ler_arquivo_xlsx(nome_arquivo):
    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
        sheet = workbook.active
        linhas_convertidas = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            linha_lista = list(row)
            linhas_convertidas.append(linha_lista)

        return linhas_convertidas

    except Exception as e:
        print("Ocorreu um erro:", e)
        return None
    
#transform
def formata_planilha(lista):
    nova_lista = []
    for index, vl in enumerate(lista):
            print(index)
            prob_acerto = gerar_probabilidade(vl, lista)
            lista_aux = sorted(vl);
            lista_aux.append(prob_acerto)
            nova_lista.append(lista_aux)
    
    return nova_lista

def gerar_probabilidade(lista, listaCompleta):

    tamanho_lista = len(listaCompleta)
    media_acertos = 0 

    for vl in listaCompleta:
        acertos = len(set(lista) & set(vl))
        media = acertos / 15
        media_acertos = media_acertos + media

    media_final = (media_acertos / tamanho_lista) * 100

    return media_final

#load
def gerar_arquivo_xlsx(lista):
    wb = openpyxl.Workbook() 
    sheet = wb.active 

    lista_cabecalho = ['B1','B2','B3','B4','B5','B6','B7','B8','B9','B10','B11',
                       'B12','B13','B14','B15','PROBABILIDADE']
    
    for i, cb in enumerate(lista_cabecalho):
        c1 = sheet.cell(row = 1, column = i+1) 
        c1.value = cb

    for index, vl in enumerate(lista):
        for idx, subl in enumerate(vl):
            c1 = sheet.cell(row = index+2, column = idx+1) 
            c1.value = subl
    
    
    wb.save("loto_load.xlsx") 


#extract
nome_arquivo = "loto.xlsx"
lista_jogos = ler_arquivo_xlsx(nome_arquivo)

#transform
lista_jogos = formata_planilha(lista_jogos)

#load
gerar_arquivo_xlsx(lista_jogos)

